import logging, re
import openpyxl
from odoo import api, fields, models, tools, _
from odoo.exceptions import UserError, ValidationError
_logger = logging.getLogger(__name__)

LINER_RTIO_CODE_KEY_PARAMS = ['material_type', 'material_grade', 'profile', 'height', 'width', 'thickness', 'no_of_fixings',
                              'fixing_specs', 'hole_ids', 'weld_overlay', 'chf_ids', 'cut_ids',
                              'cut_back_ids', 'cut_out01', 'hole01', 'wave01', 'ar_bar', 'as_drawn_opp_hand']


class StandardLinerWithDrawing(models.Model):
    _name = "standard.liner_with_drawing"
    _description = "Standard Liner With Drawing"
    _order = "standard_liner"
    _rec_name = 'standard_liner'

    standard_liner = fields.Char("Standard Liner")
    drawing = fields.Char("Drawing")
    _sql_constraints = [
        ('standard_liner_with_drawing_uniq', 'unique(standard_liner,drawing)',
         'Standard Liner With Drawing are already added.'),
    ]


class RioMaterialGradeSpecs(models.Model):
    _name = "rio.grade_specs"
    _description = "Material Grade Specs"
    _order = "name"
    _rec_name = 'name'

    name = fields.Char("Material Grade Specs")
    material_grade_id = fields.Many2one('rio.material_grade', 'Material Grade')
    _sql_constraints = [
        ('grade_specs_uniq', 'unique(name)',
         'This material grade spec is already added.'),
    ]


class RioMaterialGrade(models.Model):
    _name = "rio.material_grade"
    _description = "Material Grade"
    _order = "name"
    _rec_name = 'name'

    name = fields.Char("Material Grade")
    description = fields.Char("Description")
    _sql_constraints = [
        ('material_grade_uniq', 'unique(name)',
         'This material grade is already added.'),
    ]


class RioPu(models.Model):
    _name = "rio.pu"
    _description = "Production Unit"
    _order = "name"
    _rec_name = 'name'

    name = fields.Char("Production Unit")
    description = fields.Char("Description")
    site = fields.Many2one('rio.site')

    def import_pu_data(self, dms_id):
        dms_file = self.env['dms.file'].browse(dms_id)
        filename = dms_file.copy_file_to_temp_dir(dms_file.get_dms_file_path(), dms_file.name)
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active
        headers = [cell for cell in sheet.iter_rows(min_row=1, max_row=1, values_only=True)]
        headers = list(headers[0])
        for row in sheet.iter_rows(min_row=2):
            row_data = {headers[col_index]: cell.value for col_index, cell in enumerate(row)}
            res = self.search([('name', '=', row_data[headers[1]].strip())])
            if res:
                res.write({'description': row_data[headers[2]].strip()})
                continue
            res_site = self.env['rio.site'].search([('name', '=', row_data[headers[0]].strip())])
            if not res_site:
                raise UserError(f"error in line -> Site not available: {row_data[headers[0]]}")
            self.create(
                {'site': res_site[0].id, 'name': row_data[headers[1]].strip(),
                 'description': row_data[headers[2]].strip()})

    _sql_constraints = [
        ('pu_uniq', 'unique(name, site)',
         'This pu is already added.'),
    ]


class RioFloc(models.Model):
    _name = "rio.flocs"
    _description = "Rio FLOCs"
    _order = "name"
    _rec_name = 'name'

    name = fields.Char("FLOC")
    pu_id = fields.Many2one('rio.pu', 'PU')
    description = fields.Char("Description")
    criticality = fields.Char('Criticality')
    process_area = fields.Many2one('rio.process_area', string='Plant Area')
    asset_type = fields.Many2one('rio.asset_type', string='Asset Class')

    _sql_constraints = [
        ('floc_uniq', 'unique(name, pu_id)',
         'This Floc is already added.'),
    ]

    def import_flocs_data(self, dms_id):
        dms_file = self.env['dms.file'].browse(dms_id)
        filename = dms_file.copy_file_to_temp_dir(dms_file.get_dms_file_path(), dms_file.name)
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active
        headers = [cell for cell in sheet.iter_rows(min_row=1, max_row=1, values_only=True)]
        headers = list(headers[0])

        for row in sheet.iter_rows(min_row=2):
            vals = {}
            row_data = {headers[col_index]: cell.value for col_index, cell in enumerate(row)}
            res = self.search([('name', '=', row_data[headers[1]].strip())])

            if row_data[headers[2]]:
                vals['description'] = str(row_data[headers[2]]).strip() if row_data[headers[3]] else ''
            if row_data[headers[3]]:
                vals['criticality'] = str(row_data[headers[3]]).strip() if row_data[headers[3]] else ''

            if row_data[headers[4]]:
                process_area = self.env['rio.process_area'].search([('name', '=', row_data[headers[4]].strip())])
                if not process_area:
                    raise UserError(f"Process Area not available: {row_data[headers[4]]}")
                vals['process_area'] = process_area[0].id
            if row_data[headers[5]]:
                asset_type = self.env['rio.asset_type'].search([('name', '=', row_data[headers[5]].strip())])
                if not asset_type:
                    raise UserError(f"Asset Type not available: {row_data[headers[5]]}")
                vals['asset_type'] = asset_type[0].id

            if res:
                res.write(vals)
                continue

            pu = self.env['rio.pu'].search([('name', '=', row_data[headers[0]].strip())])
            if not pu:
                raise UserError(f"error in line -> pu not available: {row_data[headers[0]]}")
            vals['pu_id'] = pu[0].id
            vals['name'] = row_data[headers[1]].strip()
            self.create(vals)


class RioSites(models.Model):
    _name = "rio.site"
    _description = "Rio Site"
    _rec_name = "code"
    _order = "code,name"

    name = fields.Char("Name", required=True)
    code = fields.Char("Code")
    short_text = fields.Char("Short Text")
    description = fields.Char("Description")
    global_site = fields.Boolean("Is Global Site")

    _sql_constraints = [
        ('rio_site_name_uniq', 'unique(code)',
         'This Rio Site is already added.'),
    ]


class LinerMaterialType(models.Model):
    _name = "material.type"
    _description = "Material Type"
    _rec_name = "name"
    _order = "name"

    name = fields.Char("Name")
    code = fields.Char("Code")
    short_code = fields.Char("Short Code")
    description = fields.Char("Description")
    _sql_constraints = [
        ('material_type_uniq', 'unique(name)',
         'This material type is already added.'),
    ]


class LinerMaterialTypeName(models.Model):
    _name = "material.type_name"
    _description = "Material Commercial Name"
    _rec_name = "name"
    _order = "name"

    name = fields.Char("Name")
    code = fields.Char("Code")
    material_grade = fields.Many2one("rio.material_grade", string='Material Grade')
    description = fields.Char("Description")
    _sql_constraints = [
        ('material_type_name_uniq', 'unique(name)',
         'This material type name is already added.'),
    ]


class LinerType(models.Model):
    _name = "liner.type"
    _description = "Liner Type"
    _rec = "name"
    _order = "name"

    name = fields.Char("Name")
    code = fields.Char("Code")
    description = fields.Char("Description")
    _sql_constraints = [
        ('liner_type_name', 'unique(name)',
         'This Liner type name is already added.'),
    ]


class LinerProfile(models.Model):
    _name = "liner.profile"
    _description = "Liner Profile"
    _rec_name = "name"
    _order = "name"

    name = fields.Char("Name")
    code = fields.Char("Short Code")
    description = fields.Char("Description")
    _sql_constraints = [
        ('liner_profile_uniq', 'unique(name)',
         'This liner profile is already added.'),
    ]


class FixingType(models.Model):
    _name = "fixing.type"
    _description = "Fixing Type"
    _rec_name = "name"
    _order = "name"

    name = fields.Char("Name")
    code = fields.Char("Short Code")
    description = fields.Char("Description")
    _sql_constraints = [
        ('fixing_type_name_uniq', 'unique(name)',
         'This fixing type name is already added.'),
    ]


class ProductCUTBack(models.Model):
    _name = "product.cut_back"
    _description = "Product CUT Back"

    sequence = fields.Integer(string="Sr No")
    cut_back_value = fields.Char("CUT Back Dimensions")
    product_id_cut_back = fields.Many2one('product.liner_taxonomy')

    @api.onchange('cut_back_value')
    def _onchange_cut_back_value(self):
        if self.cut_back_value:
            self.cut_back_value = lowercase_x_only(self.cut_back_value)


def split_and_shift(input_string):
    pattern = r"(\d+x\d+)\|(\d+x\d+)"
    match = re.match(pattern, input_string)
    if match:
        part1, part2 = input_string.split('|')
        part1_number = int(part1.split('x')[0])
        part2_number = int(part2.split('x')[0])
        if part2_number > part1_number:
            return f"{part2}|{part1}"
        else:
            return input_string
    return input_string


class ProductCUT(models.Model):
    _name = "product.cut"
    _description = "Product CUT"

    sequence = fields.Integer(string="Sr No")
    cut_value = fields.Char("CUT Dimensions")

    product_id_cut = fields.Many2one('product.liner_taxonomy')

    @api.onchange('cut_value')
    def _onchange_cut_value(self):
        if self.cut_value:
            cut_val = lowercase_x_only(self.cut_value)
            self.cut_value = split_and_shift(cut_val)

    def split_and_shift_schedule(self):
        liner_ids = []
        for cut in self.env['product.cut'].search([]):
            cut_org = cut.cut_value
            if not cut_org:
                continue
            cut_val = lowercase_x_only(cut.cut_value)
            cut_value = split_and_shift(cut_val)
            if cut_org != cut_value:
                liner_ids.append(cut.product_id_cut.wearlinerid)
                cut.write({'cut_value': cut_value})
        _logger.info(f"{liner_ids}")


class ProductCHF(models.Model):
    _name = "product.chf"
    _description = "Product CHF"

    sequence = fields.Integer(string="Sr No")
    chf_value = fields.Char("Chamfer Dimensions")
    chamfer_configuration = fields.Selection([('Far', 'Far'),
                                     ('Near', 'Near')], string="Configuration")

    chamfer_position = fields.Selection([('Right', 'Right'),('Left', 'Left'),
                                     ('Top', 'Top'),('Bottom', 'Bottom')], string="Position")

    product_id_chf = fields.Many2one('product.liner_taxonomy')

    @api.onchange('chf_value')
    def _onchange_chf_value(self):
        if self.chf_value:
            self.chf_value = lowercase_x_only(self.chf_value)


def lowercase_x_only(input_string):
    return ''.join([char.lower() if char == 'X' else char for char in input_string])


def calculate_expression(expression):
    if expression.strip() == "?":
        return expression
    if not re.fullmatch(r'^[+-]?\d+(\.\d+)?([+-]\d+(\.\d+)?)*$', expression):
        return expression

    result = eval(expression)
    return result


class ProductHoleDimensions(models.Model):
    _name = "product.hole_dimensions"
    _description = "Product Hole Dimensions"

    sequence = fields.Integer("Fixing No.")
    h_value = fields.Char("Horizontal Dimensions in mm")
    v_value = fields.Char("Vertical Dimensions in mm")

    product_id = fields.Many2one('product.liner_taxonomy')

    @api.onchange('h_value')
    def get_h_value(self):
        if self.h_value:
            self.h_value = calculate_expression(self.h_value)

    @api.onchange('v_value')
    def get_v_value(self):
        if self.v_value:
            self.v_value = calculate_expression(self.v_value)


class RioFoundInSites(models.Model):
    _name = "rio.found_in_sites"
    _description = "Material Found In Sites"
    _order = "site"
    _rec_name = 'site'

    site = fields.Many2one('rio.site')
    material_status = fields.Boolean(string='Material Status', default=True)
    liner_taxonomy_id = fields.Many2one('product.liner_taxonomy')
    color = fields.Integer(string='Color Index', compute='_get_default_color')

    @api.depends('material_status')
    def _get_default_color(self):
        for rec in self:
            rec.color = 0
            if rec.material_status:
                rec.color = 10


class ProductLinerTaxonomy(models.Model):
    _name = "product.liner_taxonomy"
    _description = "Product Liner Taxonomy"
    _rec_name = 'wearlinerid'
    _inherit = ['mail.thread', 'mail.activity.mixin']

    data_id = fields.Integer("Data ID", help="This is used for import in xlsx")

    wearlinerid = fields.Char(string='Wear Liner ID', required=True, copy=False, readonly=1,
                              default=lambda self: self.env['ir.sequence'].next_by_code(
                                  'product.liner_taxonomy.wearlinerid'))
    duplicate = fields.Char('Duplicate', readonly=1)
    wearliner_db_internal_key = fields.Text('Wearliner DB Internal Key', readonly=1)
    rtio_liner_code = fields.Char('Rio Code', readonly=1)
    rtio_liner_code_greater_one = fields.Char('Rtio Liner Code (Count > 1)')
    same_material_with_diff_r_code = fields.Char('Same Material With Different Rio Code (Count > 1)')

    active_liners = fields.Char('Active Liners')
    pu_id = fields.Many2one('rio.pu', 'PU', required=True)
    site = fields.Many2one('rio.site', related='pu_id.site', store=True)
    floc_id = fields.Many2one('rio.flocs', 'FLOC',
                              domain="[('pu_id', '=', pu_id)]", required=True)
    pu_desc = fields.Char('PU Desc', related='pu_id.description')
    floc_desc = fields.Char('FLOC Desc', related='floc_id.description')

    dwg_linerdetail = fields.Char('DWG Liner Detail', required=True)

    itemno = fields.Char('Item No',  copy=False)
    liner_detail_qty = fields.Integer('Liner Detail Qty', copy=False)
    dwg_markingplan = fields.Char('DWG Marking Plan')
    similar_material = fields.Html("Similar Material Count", readonly=True)

    material_type = fields.Many2one("material.type", required=True)
    material_commercial_name = fields.Many2one("material.type_name", string="Material Commercial Name")

    material_grade = fields.Many2one("rio.material_grade", string='Material Grade', copy=False)

    material_grade_specs = fields.Many2one("rio.grade_specs", string='Material Grade Specs', copy=False)

    profile = fields.Many2one("liner.profile", copy=False)
    liner_type = fields.Many2one("liner.type", copy=False)

    height = fields.Integer(string='Height', copy=False)
    width = fields.Integer(string='Width', copy=False)
    thickness = fields.Integer(string='Thickness',copy=False)
    mass = fields.Float(string='Mass', copy=False)

    no_of_fixings = fields.Char('No of Fixings', copy=False)
    fixing_type = fields.Many2one("fixing.type", copy=False)

    fixing_specs = fields.Char('Fixing Specs', copy=False)
    no_of_lifting_lug = fields.Char('No of Lifting Lug', copy=False)
    lifting_lug_config = fields.Selection(string='Lifting Lug Config',
                                          selection=[('back_side', 'BACK SIDE'), ('wear_side', 'WEAR SIDE'),
                                                     ('far', 'FAR')], copy=False)
    fixingconfig = fields.Selection(string='Fixing Config',
                                    selection=[('near', 'Near'), ('far', 'FAR')], copy=False)

    hole_ids = fields.One2many('product.hole_dimensions', 'product_id', string='Hole Dimensions')

    tolerance = fields.Char('Tolerance', copy=False)
    weld_overlay = fields.Char('Weld Overlay', copy=False)
    ar_bar = fields.Selection(string='AR Bar', selection=[('yes', 'Yes'), ('no', 'No')], copy=False)
    ar_bar_position = fields.Char('AR Bar Position')

    standard_liner = fields.Char('Standard Liner', copy=False)

    chf_ids = fields.One2many('product.chf', 'product_id_chf', string='CHF')
    cut_ids = fields.One2many('product.cut', 'product_id_cut', string='CUT')
    cut_back_ids = fields.One2many('product.cut_back', 'product_id_cut_back', string='Cut Back')

    cut_out01 = fields.Char('Cut Out01', copy=False)
    hole01 = fields.Char('Hole01', copy=False)
    wave01 = fields.Char('Wave01', copy=False)
    as_drawn_opp_hand = fields.Selection(string='As Drawn/Opp Hand',
                                          selection=[('as_drawn', 'As Drawn'), ('opp_hand', 'Opp Hand')], copy=False)
    comments = fields.Char('Comments')

    sap_material = fields.Char('Material Number', default="NA", required=True)
    sap_material_description = fields.Char('Material Description', copy=False)
    material_remarks = fields.Char('Material Remarks', copy=False)

    active = fields.Boolean(string='active', default=True ,groups="rio_cyeims.group_rio_manager")

    process_area = fields.Many2one('rio.process_area', string='Process Area')
    asset_type = fields.Many2one('rio.asset_type', string='Asset Type')

    parent_relation = fields.Selection(string='Parent / Child',
                                       selection=[('parent_only', 'Parent Only'), ('parent', 'Parent'),
                                                  ('child', 'Child'), ('parent_with_drawings', 'Parent with Drawings')])

    criticality = fields.Char('Criticality')
    unspsc_code = fields.Many2one("unspsc.code", default=lambda self: self.env['unspsc.code'].search([], limit=1),
                                  string="UNSPSC Code")

    found_in_sites = fields.One2many('rio.found_in_sites', 'liner_taxonomy_id', string='Found In Sites', copy=False)
    sap_long_text = fields.Text("SAP Long Text")
    sap_short_text = fields.Text("SAP Short Text")
    is_exist_sap = fields.Boolean('Is Exist in SAP')
    dwg_ref_count = fields.Integer('DWG Ref Count')

    create_uid = fields.Many2one("res.users", string='Created by', readonly=True)
    create_date = fields.Datetime(string='Created on', readonly=True)
    write_uid = fields.Many2one("res.users", string='Last Updated by', readonly=True)
    write_date = fields.Datetime(string='Last Updated on', readonly=True)

    remarks = fields.Char('Remarks')

    proposed_material_number = fields.Char('Proposed Material Number', copy=False)
    proposed_material_desc = fields.Char('Proposed Material Desc', copy=False)

    sap_new_material = fields.Char('New Material Number', default="NA", tracking=True)

    rtiocode_repeated_multitime_same_material = fields.Char('Riocode repeated multi time same material (Count > 1)')
    multiple_material_without_na_rcode = fields.Char('Riocode with multi Material without NA (Count > 1)')
    multiple_material_with_na_rcode = fields.Char('Multiple material With NA same RIO Code (Count > 1)')

    def compute_all_filter(self):
        self.compute_riocode_repeated_multi_same_material()
        self.compute_same_riocode_multiple_material_without_na()
        self.compute_same_riocode_multiple_material_with_na()
        self.same_material_with_differnt_rio_code()
        self.compute_group_by_greater_than_one()

    def compute_riocode_repeated_multi_same_material(self):
        q = f"""SELECT rtio_liner_code, MAX(sap_material) AS sap_material, COUNT(*) AS total_rows FROM product_liner_taxonomy
                WHERE sap_material IS NOT NULL AND sap_material <> 'NA' AND rtio_liner_code NOT IN (
                SELECT DISTINCT rtio_liner_code FROM product_liner_taxonomy WHERE sap_material = 'NA' OR sap_material IS NULL)
                GROUP BY rtio_liner_code HAVING COUNT(DISTINCT sap_material) = 1 AND COUNT(*) > 1;"""
        self.env.cr.execute(q)
        result = self.env.cr.fetchall()
        rtio_codes = [r[0] for r in result]

        if len(rtio_codes) == 1:
            rtio_liner_codes_tuple = f"('{rtio_codes[0]}')"
        else:
            rtio_liner_codes_tuple = tuple(rtio_codes)

        q2 = f"UPDATE product_liner_taxonomy SET rtiocode_repeated_multitime_same_material = False;"
        self.env.cr.execute(q2)

        if rtio_liner_codes_tuple:
            q2 = f"UPDATE product_liner_taxonomy SET rtiocode_repeated_multitime_same_material = 'Riocode repeated multi same material' where rtio_liner_code IN {rtio_liner_codes_tuple};"
            self.env.cr.execute(q2)

    def compute_same_riocode_multiple_material_without_na(self):
        q = f"""SELECT rtio_liner_code, STRING_AGG(DISTINCT sap_material, ', ') AS material_combination,COUNT(*) AS count 
        FROM product_liner_taxonomy GROUP BY rtio_liner_code HAVING COUNT(DISTINCT sap_material) > 1 
        AND SUM(CASE WHEN sap_material = 'NA' THEN 1 ELSE 0 END) = 0;"""
        self.env.cr.execute(q)
        result = self.env.cr.fetchall()
        rtio_codes = [r[0] for r in result]

        if len(rtio_codes) == 1:
            rtio_liner_codes_tuple = f"('{rtio_codes[0]}')"
        else:
            rtio_liner_codes_tuple = tuple(rtio_codes)

        q2 = f"UPDATE product_liner_taxonomy SET multiple_material_without_na_rcode = False;"
        self.env.cr.execute(q2)

        if rtio_liner_codes_tuple:
            q2 = f"UPDATE product_liner_taxonomy SET multiple_material_without_na_rcode = 'Riocode with multi Material without NA' where rtio_liner_code IN {rtio_liner_codes_tuple};"
            self.env.cr.execute(q2)

    def compute_same_riocode_multiple_material_with_na(self):
        q = f"""SELECT rtio_liner_code FROM product_liner_taxonomy GROUP BY rtio_liner_code HAVING COUNT(*) > 1
            AND COUNT(DISTINCT sap_material) = 1 AND MAX(sap_material) = 'NA';"""
        self.env.cr.execute(q)
        result = self.env.cr.fetchall()
        rtio_codes = [r[0] for r in result]

        if len(rtio_codes) == 1:
            rtio_liner_codes_tuple = f"('{rtio_codes[0]}')"
        else:
            rtio_liner_codes_tuple = tuple(rtio_codes)

        q2 = f"UPDATE product_liner_taxonomy SET multiple_material_with_na_rcode = False;"
        self.env.cr.execute(q2)

        if rtio_liner_codes_tuple:
            q2 = f"UPDATE product_liner_taxonomy SET multiple_material_with_na_rcode = 'Riocode with multi Material with NA' where rtio_liner_code IN {rtio_liner_codes_tuple};"
            self.env.cr.execute(q2)

    def update_new_material_number_from_bom(self):
        bom_model = self.env['taxonomy.bom_structure']
        liner_model = self.env['product.liner_taxonomy']
        bom_rio_code_with_material = bom_model.read_group([('new_material_number', '!=', False)],
                                                          ['rio_code', 'new_material_number'], ['rio_code'])
        for bom_str in bom_rio_code_with_material:
            rtio_liner_code = bom_str.get('rio_code')
            bom = bom_model.search(bom_str.get('__domain'), limit=1)
            liners = liner_model.search([('rtio_liner_code', '=', rtio_liner_code)])
            if liners and bom.new_material_number:
                liners.sudo().write({'sap_new_material': bom.new_material_number})
            _logger.info(f"{rtio_liner_code}completed of material {bom.new_material_number}")

    def same_material_with_differnt_rio_code(self):
        q = f"""SELECT sap_material
                FROM product_liner_taxonomy
                WHERE sap_material != 'NA'
                GROUP BY sap_material
                HAVING COUNT(DISTINCT rtio_liner_code) > 1;
            """
        self.env.cr.execute(q)
        result = self.env.cr.fetchall()
        materials = [r[0] for r in result]
        if len(materials) == 1:
            materials_tuple = f"('{materials[0]}')"
        else:
            materials_tuple = tuple(materials)

        q2 = f"UPDATE product_liner_taxonomy SET same_material_with_diff_r_code = False;"
        self.env.cr.execute(q2)

        if materials_tuple:
            q2 = f"UPDATE product_liner_taxonomy SET same_material_with_diff_r_code = 'Same Material With Different RIO-Code' where sap_material IN {materials_tuple};"
            self.env.cr.execute(q2)

    def compute_group_by_greater_than_one(self):
        q = f"""SELECT rtio_liner_code, STRING_AGG(DISTINCT sap_material, ', ') AS material_combination, COUNT(*) AS count
                FROM product_liner_taxonomy
                GROUP BY rtio_liner_code
                HAVING COUNT(*) > 1  
                AND COUNT(CASE WHEN sap_material = 'NA' THEN 1 END) > 0  -- Ensures "N/A" is present
                AND COUNT(DISTINCT sap_material) > 1;"""
        self.env.cr.execute(q)
        result = self.env.cr.fetchall()
        rtio_codes = [r[0] for r in result]


        if len(rtio_codes) == 1:
            rtio_liner_codes_tuple = f"('{rtio_codes[0]}')"
        else:
            rtio_liner_codes_tuple = tuple(rtio_codes)

        q2 = f"UPDATE product_liner_taxonomy SET rtio_liner_code_greater_one = False;"
        self.env.cr.execute(q2)

        if rtio_liner_codes_tuple:
            q2 = f"UPDATE product_liner_taxonomy SET rtio_liner_code_greater_one = 'Same RIO-Code With NA' where rtio_liner_code IN {rtio_liner_codes_tuple};"
            self.env.cr.execute(q2)

    def write(self, vals):
        if not self.env.user.has_group('rio_cyeims.group_rio_liner_taxonomy_admin'):
            for record in self:
                if record.sap_new_material and record.sap_new_material.strip().upper() != "NA":
                    raise ValidationError(
                        "You cannot modify any fields once 'New Material Number' is set. Liner Taxonomy "
                        "Admin can Change Only")

        if 'liner_detail_qty' in vals.keys() and vals.get('liner_detail_qty', 0) == 0:
            raise UserError(f"Liner Details Quantity Can not be 0 .")
        res = super(ProductLinerTaxonomy, self).write(vals)
        return res

    @api.model
    def create(self, values):
        if 'liner_detail_qty' in values.keys() and values.get('liner_detail_qty', 0) == 0:
            raise UserError(f"Liner Details Quantity Can not be 0 .")
        rec = super(ProductLinerTaxonomy, self).create(values)
        return rec

    @api.onchange('standard_liner')
    def _onchange_standard_liner(self):
        if self.standard_liner:
            self.standard_liner = self.standard_liner.replace('w', 'W').replace('l', 'L')

    @api.onchange('fixing_specs')
    def _onchange_fixing_specs(self):
        if self.fixing_specs:
            self.fixing_specs = lowercase_x_only(self.fixing_specs)

    @api.onchange('cut_out01')
    def _onchange_cut_out01(self):
        if self.cut_out01:
            self.cut_out01 = lowercase_x_only(self.cut_out01)

    @api.onchange('hole01')
    def _onchange_hole01(self):
        if self.hole01:
            self.hole01 = lowercase_x_only(self.hole01)

    @api.onchange('wave01')
    def _onchange_wave01(self):
        if self.wave01:
            self.wave01 = lowercase_x_only(self.wave01)

    @api.onchange('weld_overlay')
    def get_weld_overlay(self):
        pattern = r"[0-9]{1,2}[/][0-9]{1,2}"
        if self.weld_overlay:
            if not re.match(pattern, self.weld_overlay):
                return {
                    'warning': {'title': "Data Rules",
                                'message': "DR0002: Please Enter Valid Weld Overlay in 00/00 format",
                                'type': 'notification'},
                }
            overlay_part = self.weld_overlay.split('/')
            part1 = int(overlay_part[0])
            part2 = int(overlay_part[1])
            total = part1 + part2
            self.thickness = total

    @api.onchange('no_of_fixings')
    def check_product_hole_dimensions(self):
        if not self.no_of_fixings:
            return
        fix_value = int(self.no_of_fixings)
        for record in self:
            if len(record.hole_ids) != fix_value:
                return {
                    'warning': {'title': "Data Rules",
                                'message': "DR0001: No of Fixings and Fixing Number are Mismatch",
                                'type': 'notification'},
                }

    def get_string_data(self, str):
       return str if str else ''

    def prepare_sap_material_short_and_long_text(self, object_id=None):
        material = self.browse(object_id) if object_id else self
        separator = ';'
        short_text = "LINER," + self.get_string_data(material.material_type.short_code) + separator

        if material.standard_liner:
            liner = material.standard_liner.lower()
            if not ('modified' in liner or 'type' in liner):
                short_text += 'STD'+separator

        short_text += self.get_string_data(material.profile.code) + separator
        if self.get_string_data(material.fixing_type.code) and self.get_string_data(material.no_of_fixings):
            short_text += f'{self.get_string_data(material.fixing_type.code)}x{self.get_string_data(material.no_of_fixings)}{separator}'
        short_text += f'{self.get_string_data(material.height)}x{self.get_string_data(material.width)}x{self.get_string_data(material.thickness)}mm'

        short_text = short_text.strip(separator)
        sap_long_text, dwg_count = material.get_long_text(material)
        material.write({"sap_short_text": short_text, "sap_long_text": sap_long_text, 'dwg_ref_count': dwg_count})

    def get_long_text(self, material):
        dwg_count = 0
        separator = '\n'
        long_text = f""
        std_flag = False
        if material.material_type:
            long_text += f"Material Type: {material.material_type.name}{separator}"

        if material.standard_liner:
            value = material.standard_liner.lower()
            if not ('modified' in value.lower() or 'type' in value.lower()):
                long_text += f"Standard Liners: {material.standard_liner}{separator}"
                std_flag = True

        if material.profile:
            long_text += f"Material Profile: {material.profile.name}{separator}"
        if material.height:
            long_text += f"Length: {material.height}{separator}"
        if material.width:
            long_text += f"Width: {material.width}{separator}"
        if material.thickness:
            long_text += f"Thickness: {material.thickness}{separator}"
        if material.fixing_type:
            long_text += f"Fixing Type: {material.fixing_type.name}{separator}"
        if material.no_of_fixings:
            long_text += f"Number of Fixing Points: {material.no_of_fixings}{separator}"
        if material.fixing_specs:
            long_text += f"Fixing Specification: {material.fixing_specs}{separator}"
        if material.no_of_lifting_lug:
            long_text += f"No: of lifting Lugs: {material.no_of_lifting_lug}{separator}"

        if material.tolerance:
            long_text += f"Tolerance: {material.tolerance}{separator}"
        if material.material_grade:
            long_text += f"Material Grade: {material.material_grade.name}{separator}"
        if material.weld_overlay:
            long_text += f"Weld Overlay: {material.weld_overlay}{separator}"
        if material.ar_bar:
            long_text += f"AR Bar: {material.ar_bar}{separator}"

        if material.profile.name.lower() in ['rectangle', 'square']:
            long_text += f"Special Features: No Special Features {separator}"
        if material.chf_ids:
            cut_vals = f"Chamfer: "
            for cut in material.chf_ids:
                cut_vals += f"{cut.chf_value}; "
            long_text += f"{cut_vals} {separator}"
        if material.cut_ids:
            cut_vals = f"Cut: "
            for cut in material.cut_ids:
                cut_vals += f"{cut.cut_value}; "
            long_text += f"{cut_vals} {separator}"
        if material.cut_back_ids:
            cut_vals = f"Cut Back: "
            for cut in material.cut_back_ids:
                cut_vals += f"{cut.cut_back_value}; "
            long_text += f"{cut_vals} {separator}"
        if material.cut_out01:
            long_text += f"Cut Out: {material.cut_out01}{separator}"
        if material.hole01:
            long_text += f"Hole: {material.hole01}{separator}"
        if material.wave01:
            long_text += f"Wave: {material.wave01}{separator}"
        if material.comments:
            long_text += f"Additional Comments: {material.comments}{separator}"

        if material.dwg_linerdetail:
            if std_flag:
                ref_text = f""
                value = material.standard_liner.lower()
                std_records = self.env['standard.liner_with_drawing'].search([])
                for std_liner in std_records:
                    if std_liner.standard_liner.lower() == value:
                        ref_text += f'Reference Drawing 1: {std_liner.drawing}; Item No: {std_liner.standard_liner}{separator}'
                        dwg_count += 1
                        break
            else:
                records = self.env['product.liner_taxonomy'].search([('sap_material', '=', material.sap_material),
                                                                     ('wearliner_db_internal_key', '=',
                                                                      material.wearliner_db_internal_key)])
                ref_dwg = list(
                    set([f"{rec.dwg_linerdetail}; Item No: {rec.itemno}" for rec in records if rec.dwg_linerdetail]))
                ref = 0
                ref_text = f""
                for dwg in ref_dwg:
                    ref += 1
                    ref_text += f'Reference Drawing {ref}: {dwg}{separator}'
                dwg_count = ref

            long_text += ref_text

        long_text = long_text.strip(separator)
        return long_text, dwg_count

    def prepare_sap_material_no(self, domain=[]):
        domain = domain + [('sap_material', '=', 'NA')]
        count = 0
        for material in self.search(domain):
            count = count + 1
            suffix = f"{material.material_type.code or 'NA'}"
            new_mat_code = material.rtio_liner_code.replace(suffix, "NEWMAT")
            material.write({"sap_material": new_mat_code})
            if count > 100:
                self._cr.commit()
                count = 0

    def get_found_in_sites(self):
        offset = 0
        limit = 1000
        not_all_done = True

        while not_all_done:
            records = self.env['product.liner_taxonomy'].search([('sap_material', '!=', 'NA')], limit=limit,
                                                                offset=offset)
            for record in records:
                z4191_data = self.env['z4191.sample_customized'].read_group([('material', '=', record.sap_material)],
                                                                            ['plant'], ['plant'])
                mm60_data = self.env['mm60.sample_customized'].read_group([('material', '=', record.sap_material)],
                                                                          ['plnt'], ['plnt'])
                plant_list_z4191 = list(item['plant'] for item in z4191_data)
                plant_list_mm60 = list(item['plnt'] for item in mm60_data)
                plant_list_z4191.extend(plant_list_mm60)
                plant_list = list(set(plant_list_z4191))
                sites = self.env['rio.site'].search([('code', 'in', plant_list)])
                sites_ids = []

                for site in sites:
                    vals = {'site': site.id}
                    mt_status = self.env['mb52.sample_customized'].search([('material', '=', record.sap_material),
                                                                           ('df_stor_loc_level', '=', 'X'),
                                                                           ('plant', '=', site.code)], limit=1)
                    vals['material_status'] = True
                    if mt_status:
                        vals['material_status'] = False
                    sites_ids.append(vals)

                    status = self.env['rio.found_in_sites'].search([('site', '=', site.id),
                                                                    ('liner_taxonomy_id', '=', record.id)])
                    if status:
                        status.write(vals)
                    else:
                        vals['liner_taxonomy_id'] = record.id
                        self.env['rio.found_in_sites'].create(vals)

            self._cr.commit()
            offset += limit
            _logger.info(f"{limit}completed {offset}")
            if not records:
                not_all_done = False
                _logger.info(f"Cron Job is completed {offset}")

    def update_similar_material_count(self, params):
        profile_ids = params.get('profile_ids', [])
        tolerance = params.get('tolerance', 3)
        limit = params.get('limit', 100)

        count = 0
        limit_cal = 0
        for record in self.search([('profile', 'in', profile_ids)], order='site, material_type, width'):
            thickness = record.thickness or 0
            width = record.width or 0
            height = record.height or 0
            if thickness == 0 or width == 0 or height == 0:
                continue
            params = {'record_id': record.id, 'tolerance': tolerance}
            similar_materials = self.get_similar_material_record_ids(params)
            similar_material_count = len(similar_materials)
            if similar_material_count < 2:
                continue

            action_id = self.env['ir.actions.act_window'].search([('name', '=', 'Liner Taxonomy Similar Materials')])
            base_url = self.env['ir.config_parameter'].get_param('web.base.url')
            params = f"&record_id={record.id}&tolerance={tolerance}"
            link = f"{base_url}/web#cids=1&menu_id=500&action={action_id.id}&model=product.liner_taxonomy&view_type=list{params}"
            similar_materials_url = f"<a href='{link}' target='_blank' style='color: blue;'>&nbsp;&nbsp;&nbsp;&nbsp;{similar_material_count}</a>"

            self.browse(similar_materials).write({'similar_material': similar_materials_url})
            count = count + 1
            limit_cal = limit_cal + 1
            if count > 100:
                self._cr.commit()
                count = 0
            if limit and limit_cal > limit:
                break
        self._cr.commit()
        return True

    def get_rtio_multiple_mtr_records(self, site_id):
        domain = [('site', '=', site_id)]
        reo_site = self.env['rio.site'].browse(site_id)
        if reo_site.global_site:
            domain = []

        rtio_multiple_mtr = {}
        records = []
        for record in super().search(domain):
            sap_material = record.sap_material if record.sap_material.strip() and record.sap_material != 'NA' else False
            if sap_material:
                if record.rtio_liner_code in rtio_multiple_mtr:
                    if sap_material not in rtio_multiple_mtr[record.rtio_liner_code][0]:
                        rtio_multiple_mtr[record.rtio_liner_code] = (rtio_multiple_mtr[record.rtio_liner_code][0] + [
                            sap_material], rtio_multiple_mtr[record.rtio_liner_code][1] + [record.id])
                else:
                    rtio_multiple_mtr[record.rtio_liner_code] = ([sap_material], [record.id])

        count_rtio_multiple_mtr = 0
        for k, v in rtio_multiple_mtr.items():
            if len(v[0]) > 1:
                count_rtio_multiple_mtr = count_rtio_multiple_mtr + 1
                records = records + v[1]
        return records

    @api.model
    def search_read(self, domain=None, fields=None, offset=0, limit=None, order=None):
        res = super().search_read(domain, fields, offset, limit, order)
        return res

    @api.model
    def web_read_group(self, domain, fields, groupby, limit=None, offset=0, orderby=False,
                       lazy=True, expand=False, expand_limit=None, expand_orderby=False):
        action = self.env.context.get('action', False)
        if action == 'rtio_multiple_mtr' and not domain:
            params = self.env.context.get('params', False)
            if params and params.get("site_id", False):
                records = self.get_rtio_multiple_mtr_records(params["site_id"])
                domain = [('id', 'in', records)]

        elif action == 'similar_materials' and not domain:
            params = self.env.context.get('params', False)
            unique_records = self.get_similar_material_record_ids(params)
            if unique_records:
                domain = [('id', 'in', unique_records)]

        return super().web_read_group(domain, fields, groupby, limit=limit, offset=offset, orderby=orderby, lazy=lazy,
                                      expand=expand, expand_limit=expand_limit, expand_orderby=expand_orderby)

    @api.model
    def read_group(self, domain, fields, groupby, offset=0, limit=None, orderby=False, lazy=True):
        action = self.env.context.get('action', False)
        if action == 'rtio_multiple_mtr' and not domain:
            params = self.env.context.get('params', False)
            if params and params.get("site_id", False):
                records = self.get_rtio_multiple_mtr_records(params["site_id"])
                domain = [('id', 'in', records)]

        elif action == 'similar_materials' and not domain:
            params = self.env.context.get('params', False)
            unique_records = self.get_similar_material_record_ids(params)
            if unique_records:
                domain = [('id', 'in', unique_records)]
        return super().read_group(domain, fields, groupby, offset=offset, limit=limit, orderby=orderby, lazy=lazy)

    def get_similar_material_record_ids(self, params):
        if params and params.get("record_id", False):
            tolerance = params.get("tolerance", 3)
            record = self.browse(params.get("record_id", False))
            width = [('width', '>', record.width - tolerance), ('width', '<', record.width + tolerance)]
            height = [('height', '>', record.height - tolerance), ('height', '<', record.height + tolerance)]
            thickness = [('thickness', '>', record.thickness - tolerance),
                         ('thickness', '<', record.thickness + tolerance)]
            search_params = [('material_type', '=', record.material_type.id),
                             ('profile', '=', record.profile.id)] + width + height + thickness

            similar_material = self.search(search_params)
            unique_rio_code = {record.rtio_liner_code: record.id for record in similar_material}
            unique_records = list(unique_rio_code.values())
            return unique_records
        return []

    def search(self, args, offset=0, limit=None, order=None, count=False):
        action = self.env.context.get('action', False)
        if action == 'rtio_multiple_mtr' and not args:
            params = self.env.context.get('params', False)
            if params and params.get("site_id", False):
                records = self.get_rtio_multiple_mtr_records(params["site_id"])
                return self.browse(records)

        if action == 'similar_materials' and not args:
            params = self.env.context.get('params', False)
            unique_records = self.get_similar_material_record_ids(params)
            if unique_records:
                return self.search([('id', 'in', unique_records)])

        res = super().search(args, offset, limit, order, count)
        return res

    def update_rtio_liner_code(self, force_update=False, record=False):
        record = record or self
        if not force_update and record.rtio_liner_code:
            _logger.info(f"RTIO Liner Code: {record.rtio_liner_code} already assigned to this record")
            return True
        rtio_liner_code = False
        if not record.wearliner_db_internal_key:
            _logger.info(f"Liner key is not updated for this record")
            return True
        if not record.rtio_liner_code or force_update:
            record_with_rtio_liner_code = self.search(
                [('wearliner_db_internal_key', '=', record.wearliner_db_internal_key),
                 ('rtio_liner_code', '!=', False), ('id', '!=', record.id)])
            if record_with_rtio_liner_code:
                rtio_liner_code = record_with_rtio_liner_code[0].rtio_liner_code

        if not rtio_liner_code:
            if record.rtio_liner_code:
                res = self.search([('rtio_liner_code', '=', record.rtio_liner_code), ('id', '!=', record.id)])
                if res:
                    rtio_liner_code = f"{record.material_type.code or 'NA'}{self.env['ir.sequence'].next_by_code('product.liner_taxonomy.rtio_liner_code')}"
                else:
                    rtio_liner_code = record.rtio_liner_code
            else:
                rtio_liner_code = f"{record.material_type.code or 'NA'}{self.env['ir.sequence'].next_by_code('product.liner_taxonomy.rtio_liner_code')}"
        code_prefix = rtio_liner_code[:2]
        if record.material_type.code != code_prefix:
            rtio_liner_code = f"{record.material_type.code or 'NA'}{rtio_liner_code[2:]}"

        record.write({'rtio_liner_code': rtio_liner_code})
        count = self.search_count([('rtio_liner_code', '=', rtio_liner_code)])
        record.write({'duplicate': 'Yes' if count > 1 else 'No'})
        record.prepare_sap_material_short_and_long_text()

    def update_record_rtio_liner_key(self, args):
        records = True
        limit = args.get('limit', 100)
        force_update = args.get('force_update', False)
        domain = []
        site = args.get('site_id', False)
        if site:
            domain.append(('site', '=', site))

        if not force_update:
            domain = [('wearliner_db_internal_key', '=', False)]
        offset = 0
        while True:

            records = self.search(domain, offset=offset, limit=limit)
            if not records:
                break
            for record in records:
                record.action_calculate_record_rtio_liner_code(force_update)
            offset += limit
            self._cr.commit()
            _logger.info(f"{offset} -- {limit} RIO-Code updated")

    def action_calculate_record_rtio_liner_code(self, force_update=True, record=False):
        key = self.get_key(record or self)
        self.update_rtio_liner_code(force_update, record or self)
        return True

    def get_key(self, record):
        domain = [('model_id', '=', 'product.liner_taxonomy')]
        ir_fields = [field_id for field_id in self.env['ir.model.fields'].sudo().search(domain)]
        ir_fields = {field_id.name: field_id for field_id in ir_fields}
        key_value_pairs = []
        for attr in LINER_RTIO_CODE_KEY_PARAMS:
            if record[attr]:
                ir_field = ir_fields[attr]
                name = ir_field.display_name.replace('(product.liner_taxonomy)', '').strip()
                if ir_field.ttype == "many2one":
                    value = record[attr].name
                elif ir_field.ttype == "one2many":
                    value = record[attr]
                    if ir_field.relation == 'product.hole_dimensions':
                        values = []
                        sequence = 0
                        for hole in value:
                            sequence = sequence + 1
                            h_value = hole.h_value if hole.h_value else ""
                            v_value = hole.v_value if hole.v_value else ""
                            values.append(f"(H{sequence}: {h_value}, V{sequence}: {v_value})")
                        if values:
                            value = ", ".join(values)
                    elif ir_field.relation == 'product.cut_back':
                        values = []
                        sequence = 0
                        for hole in value:
                            sequence = sequence + 1
                            cut_back_value = hole.cut_back_value if hole.cut_back_value else ""
                            values.append(f"Cut_Back{sequence}: {cut_back_value}")
                        if values:
                            value = ", ".join(values)
                            value = f"({value})"

                    elif ir_field.relation == 'product.cut':
                        values = []
                        sequence = 0
                        for hole in value:
                            sequence = sequence + 1
                            cut_value = hole.cut_value if hole.cut_value else ""
                            values.append(f"Cut{sequence}: {cut_value}")
                        if values:
                            value = ", ".join(values)
                            value = f"({value})"

                    elif ir_field.relation == 'product.chf':
                        values = []
                        sequence = 0
                        for hole in value:
                            sequence = sequence + 1
                            chf_value = hole.chf_value if hole.chf_value else ""
                            values.append(f"Chf{sequence}: {chf_value}")
                        if values:
                            value = ", ".join(values)
                            value = f"({value})"
                elif ir_field.ttype == "selection":
                    var_dict = dict(self._fields[ir_field.name].selection)
                    value = var_dict[record[attr]]
                else:
                    value = record[attr]
                key_value_pairs.append(f"{name}: {value}")
        if key_value_pairs:
            key = ", ".join(key_value_pairs)
        else:
            key = False
        record.write({"wearliner_db_internal_key": key})
        return key
