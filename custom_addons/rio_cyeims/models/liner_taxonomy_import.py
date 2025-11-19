import logging
import re
import openpyxl

from odoo import api, fields, models, tools, exceptions, _

_logger = logging.getLogger(__name__)

import_fields = ['data id', 'active liners', 'site', 'pu', 'floc', 'dwg liner detail', 'item no',
                 'liner detail qty', 'liner type', 'material grade specs',
                 'dwg marking plan', 'material type', 'material commercial name', 'profile', 'height',
                 'width', 'thickness', 'mass', 'no of fixings', 'fixing type', 'fixing specs', 'no of lifting lug',
                 'lifting lug config', 'fixing config', 'h01', 'h02', 'h03', 'h04', 'h05', 'h06', 'h07', 'h08', 'h09',
                 'h10', 'v01', 'v02', 'v03', 'v04', 'v05', 'v06', 'v07', 'v08', 'v09', 'v10', 'tolerance',
                 'weld overlay', 'ar bar', 'standard liner', 'chaf01', 'chaf02', 'chaf03', 'cut01', 'cut02', 'cut03',
                 'cut back01', 'cut back02', 'cut back03', 'cut out01', 'hole01', 'wave01', 'as drawn/opp hand',
                 'comments', 'material number', 'material description', 'material remarks']
hole_dim = [('h01', 'v01'), ('h02', 'v02'), ('h03', 'v03'), ('h04', 'v04'), ('h05', 'v05'), ('h06', 'v06'),
            ('h07', 'v07'), ('h08', 'v08'), ('h09', 'v09'), ('h10', 'v10')]

CHF_DIM = ['chaf01', 'chaf02', 'chaf03', 'chaf04', 'chaf05', 'chaf06', 'chaf07', 'chaf08', 'chaf09', 'chaf10']
CUT_DIM = ['cut01', 'cut02', 'cut03', 'cut04', 'cut05', 'cut06', 'cut07', 'cut08', 'cut09', 'cut10']
CUT_BACK_DIM = ['cut back01', 'cut back02', 'cut_back03', 'cut back04', 'cut back05', 'cut back06', 'cut back07',
                'cut back08', 'cut back09', 'cut back10']


class ProductLinerTaxonomyImport(models.Model):
    _inherit = 'product.liner_taxonomy'

    def read_xlsx_to_dict(self, filename):
        domain = [('model_id', '=', 'product.liner_taxonomy')]
        ir_fields = [field_id for field_id in self.env['ir.model.fields'].sudo().search(domain)]
        ir_fields = {field_id.display_name.replace('(product.liner_taxonomy)', '').strip(): field_id.name for field_id
                     in
                     ir_fields}
        ir_fields = {k.lower(): v for k, v in ir_fields.items() if k.lower() in import_fields}
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active

        headers = [cell for cell in sheet.iter_rows(min_row=1, max_row=1, values_only=True)]
        headers = list(headers[0])
        headers = [x.replace('_', ' ').replace('\n', '').title() if x else None for x in headers]
        row_index = 0
        records = 0

        def normalise(v):
            if isinstance(v, str):
                return v.replace("\xa0", " ").strip()
            return v

        rio_site = {}
        rio_material_type = {}
        rio_material_type_name = {}
        rio_liner_type = {}
        rio_liner_profile = {}
        rio_fixing_type = {}
        material_grade_specs_type = {}
        pu_type = {}
        flocs = {}
        logs = ''

        for row in sheet.iter_rows(min_row=2):
            row_data = {headers[col_index]: cell.value for col_index, cell in enumerate(row)}
            row_data = {k.lower(): normalise(v) for k, v in row_data.items() if k and k.lower() in import_fields}
            vals = {ir_fields[k]: v for k, v in row_data.items() if k in ir_fields.keys()}
            data_id = vals['data_id']
            res = self.search([('data_id', '=', data_id), ('site', '=', vals['site'])])
            if not res:
                lifting_lug_config = vals.get("lifting_lug_config", False)
                if lifting_lug_config:
                    vals['lifting_lug_config'] = lifting_lug_config.lower().replace(' ', '_')

                ar_bar = vals.get("ar_bar", False)
                if ar_bar:
                    vals['ar_bar'] = ar_bar.lower().replace(' ', '_')

                as_drawn_opp_hand = vals.get("as_drawn_opp_hand", False)
                if as_drawn_opp_hand:
                    vals['as_drawn_opp_hand'] = as_drawn_opp_hand.lower().replace(' ', '_')

                material_type = vals.get("material_type", False)
                if material_type:
                    material_type_id = rio_material_type.get(material_type, False)
                    if not material_type_id:
                        res = self.env["material.type"].search(
                            ['|', ('name', '=ilike', material_type), ('description', '=ilike', material_type)])
                        if not res:
                            self._cr.commit()
                            raise exceptions.UserError(f"Material Type \"{material_type}\" is not available")
                        material_type_id = res[0].id
                        rio_material_type[material_type] = material_type_id
                    vals['material_type'] = material_type_id

                material_type_name = vals.get("material_commercial_name", False)
                if material_type_name:
                    material_type_name_id = rio_material_type_name.get(material_type_name, False)
                    if not material_type_name_id:
                        res = self.env["material.type_name"].search([('name', '=ilike', material_type_name)])
                        if not res:
                            self._cr.commit()
                            raise exceptions.UserError(
                                f"Material Commercial Name \"{material_type_name}\" is not available")
                        material_type_name_id = res[0].id
                        rio_material_type_name[material_type_name] = material_type_name_id
                    vals['material_commercial_name'] = material_type_name_id

                liner_type = vals.get("liner_type", False)
                if liner_type:
                    liner_type_id = rio_liner_type.get(liner_type, False)
                    if not liner_type_id:
                        res = self.env["liner.type"].search([('name', '=ilike', liner_type)])
                        if not res:
                            self._cr.commit()
                            raise exceptions.UserError(f"Liner Type \"{liner_type}\" is not available")
                        liner_type_id = res[0].id
                        rio_liner_type[liner_type] = liner_type_id
                    vals['liner_type'] = liner_type_id

                liner_profile = vals.get("profile", False)
                if liner_profile:
                    liner_profile_id = rio_liner_profile.get(liner_profile, False)
                    if not liner_profile_id:
                        res = self.env["liner.profile"].search([('name', '=ilike', liner_profile)])
                        if not res:
                            self._cr.commit()
                            raise exceptions.UserError(f"Liner Profile \"{liner_profile}\" is not available")
                        liner_profile_id = res[0].id
                        rio_liner_profile[liner_profile] = liner_profile_id
                    vals['profile'] = liner_profile_id

                fixing_type = vals.get("fixing_type", False)
                if fixing_type:
                    fixing_type_id = rio_fixing_type.get(fixing_type, False)
                    if not fixing_type_id:
                        res = self.env["fixing.type"].search([('name', '=ilike', fixing_type)])
                        if not res:
                            self._cr.commit()
                            raise exceptions.UserError(f"Fixing Type \"{fixing_type}\" is not available")
                        fixing_type_id = res[0].id
                        rio_fixing_type[fixing_type] = fixing_type_id
                    vals['fixing_type'] = fixing_type_id

                material_grade_specs = vals.get("material_grade_specs", False)
                if material_grade_specs:
                    material_grade_specs_id = material_grade_specs_type.get(material_grade_specs, False)
                    if not material_grade_specs_id:
                        res = self.env["rio.grade_specs"].search([('name', '=ilike', material_grade_specs)])
                        if not res:
                            self._cr.commit()
                            raise exceptions.UserError(
                                f"Material Grade Specs \"{material_grade_specs}\" is not available")
                        material_grade_specs_type[material_grade_specs] = res[0].id
                    vals['material_grade_specs'] = material_grade_specs_type[material_grade_specs]

                pu = vals.get("pu_id", False)
                if pu:
                    pu_id = pu_type.get(pu, False)
                    if not pu_id:
                        res = self.env["rio.pu"].search([('name', '=ilike', pu)])
                        if not res:
                            self._cr.commit()
                            raise exceptions.UserError(
                                f"Production Unit(PU) \"{pu}\" is not available")
                        pu_type[pu] = res[0].id
                        pu_id = pu_type[pu]
                    vals['pu_id'] = pu_id

                floc = vals.get("floc_id", False)
                if floc:
                    floc_id = flocs.get(floc, False)
                    if not floc_id:
                        res = self.env["rio.flocs"].search([('name', '=ilike', floc)])
                        if not res:
                            self._cr.commit()
                            raise exceptions.UserError(
                                f"FLOC \"{floc}\" is not available")
                        flocs[floc] = res[0].id
                        floc_id = flocs[floc]
                    vals['floc_id'] = floc_id
                del vals['site']
                liner = self.search([('pu_id', '=', vals.get('pu_id')), ('floc_id', '=', vals.get('floc_id')),
                                   ('itemno', '=', vals.get('itemno')),
                                   ('dwg_linerdetail', '=', vals.get('dwg_linerdetail')),
                                   ('as_drawn_opp_hand', '=', vals.get('as_drawn_opp_hand'))])
                if len(liner) > 1:
                    _logger.info(f"{vals.get('itemno')};{vals.get('dwg_linerdetail')};{len(liner)}")
                    logs += f"Item No = {vals.get('itemno')}; dwg_linerdetail := {vals.get('dwg_linerdetail')}; counts = {len(liner)} <br>"
                    continue

                if liner:
                    liner.write(vals)
                    logs += f"Re-Write Item No = {vals.get('itemno')}; dwg_linerdetail := {vals.get('dwg_linerdetail')}; counts = {len(liner)} <br>"
                else:
                    liner = self.create(vals)

                self.add_hole_dim(row_data, liner)
                self.add_chf(row_data, liner)
                self.add_cut(row_data, liner)
                self.add_cut_back(row_data, liner)
                row_index = row_index + 1
            if row_index > 100:
                records = records + 100
                self._cr.commit()
                _logger.info(f"updated {records} records. logs -->{logs}")
                row_index = 0
        return logs

    def add_cut_back(self, row_data, product):
        if product.cut_back_ids:
            product.cut_back_ids.unlink()
        values = []
        sequence = 0
        for dim in CUT_BACK_DIM:
            cut_back_value = False
            if dim in row_data.keys():
                cut_back_value = row_data[dim]
            if cut_back_value:
                sequence = sequence + 1
                values = values + [
                    {'sequence': sequence, 'cut_back_value': cut_back_value, 'product_id_cut_back': product.id}]
        if values:
            self.env['product.cut_back'].create(values)

    def add_cut(self, row_data, product):
        if product.cut_ids:
            product.cut_ids.unlink()
        values = []
        sequence = 0
        for dim in CUT_DIM:
            cut_value = False
            if dim in row_data.keys():
                cut_value = row_data[dim]
            if cut_value:
                sequence = sequence + 1
                values = values + [{'sequence': sequence, 'cut_value': cut_value, 'product_id_cut': product.id}]
        if values:
            self.env['product.cut'].create(values)

    def add_chf(self, row_data, product):
        if product.chf_ids:
            product.chf_ids.unlink()
        values = []
        sequence = 0
        for dim in CHF_DIM:
            chf_value = False
            if dim in row_data.keys():
                chf_value = row_data[dim]
            if chf_value:
                sequence = sequence + 1
                values = values + [{'sequence': sequence, 'chf_value': chf_value, 'product_id_chf': product.id}]
        if values:
            self.env['product.chf'].create(values)

    def add_hole_dim(self, row_data, product):
        if product.hole_ids:
            product.hole_ids.unlink()
        values = []
        sequence = 0
        for dim in hole_dim:
            h, v = dim
            h_val = v_val = False
            if h in row_data.keys():
                h_val = row_data[h]
            if v in row_data.keys():
                v_val = row_data[v]
            if h_val or v_val:
                sequence = sequence + 1
                values = values + [{'sequence': sequence, 'h_value': h_val, 'v_value': v_val, 'product_id': product.id}]
        if values:
            self.env['product.hole_dimensions'].create(values)
