import logging

from odoo import api, fields, models, tools, _
import openpyxl, io, base64

_logger = logging.getLogger(__name__)

NEW_EXISTING_MATERIAL = [('new_material', 'Create New Material in SAP'),
 ('existing_no_match', 'Material Found in SAP But Not in Taxonomy'),
 ('existing_flocs_mismatch',  'Material Found in SAP Under A Different FLOC in Taxonomy'),
 ('add', 'Assign Existing SAP Material to The FLOC'),
 ('undefined', 'Undefined'),
 ('match', 'Material Matches both in Taxonomy and SAP'),
 ('same_riocode_material', 'Assign Existing Material having same Riocode')]

class SparesFinder(models.Model):
    _name = "spares.finder_report"
    _description = "Spares Finder Report"

    name = fields.Char("Name", readonly=True)
    re_update_bom = fields.Boolean("Re-Upload BoM")
    spares_finder = fields.Many2many('ir.attachment', 'rep_data', 'doc_id1', 'attach_id3',
                                    string="Spares Finder")
    is_system = fields.Boolean("is System", default=False)

    def action_re_upload_spares_bom(self):
        if not self.spares_finder:
            return

        excel_bytes = base64.b64decode(self.spares_finder.datas)
        excel_file = io.BytesIO(excel_bytes)

        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active

        headers = [cell for cell in sheet.iter_rows(min_row=1, max_row=1, values_only=True)]
        headers = list(headers[0])
        headers = [x.replace(' ', '_').replace('\n', '').title() if x else None for x in headers]

        header_name = ['Material_Number', 'Riocode']
        flag = all(x == y for x, y in zip(headers, header_name))

        if not flag:
            return

        for row in sheet.iter_rows(min_row=2):
            row_data = {headers[col_index]: cell.value for col_index, cell in enumerate(row)}
            bom_str_objs = self.env['taxonomy.bom_structure'].search([('rio_code', '=', row_data['Riocode'])
                                                                      , ('material_number', 'ilike', 'NEWMAT%')])

            if bom_str_objs:
                bom_str_objs.write({'new_material_number':row_data['Material_Number']})


class TaxonomyBOMStructure(models.Model):
    _name = "taxonomy.bom_structure"
    _description = "Taxonomy BOM Structure"
    _rec_name = 'package_unit'

    package_unit = fields.Char('Package Unit')
    site = fields.Many2one('rio.site', readonly=True)
    functional_location = fields.Char('Functional Location')
    floc_description = fields.Char('Flocs Description')
    ct = fields.Char('CT')
    pma = fields.Char('PMA')
    parent_pma = fields.Char('Parent PMA')
    pma_description = fields.Char('PMA Description')
    material_number = fields.Char('Material Number')
    new_material_number = fields.Char('New Material Number', tracking=True)
    material_description = fields.Char('Material Description')
    material_qty = fields.Char('Material Quantity')
    material_unit = fields.Char('Unit', deafult="EA")
    new_existing_material = fields.Selection(NEW_EXISTING_MATERIAL, string='New/Existing Material')

    action = fields.Char('Action')
    material_where_used = fields.Char('Material Where Used')
    comment = fields.Char('Comment')
    rio_code = fields.Char('Rio Code')
    liner_id = fields.Char('Liner ID')
    std_liner = fields.Char('Std Liner')
    item_no = fields.Char('Item No.')
    #layout_drawing_no = fields.Char('Layout Drawing No.')
    layout_drawing_no = fields.Char('DWG Marking Plan')
    #detail_drawing = fields.Char('Detail Drawing')
    detail_drawing = fields.Char('DWG Liner Detail')
    sap_long_text = fields.Text("SAP Long Text")
    sap_short_text = fields.Text("SAP Short Text")

    plant = fields.Char('Plant')  # MB51 – Plant wise Material availability & Goods Movement
    last_wo_used = fields.Char('Last WO used')
    goods_movement = fields.Char('Goods Movement')

    current_stock = fields.Char('Current Stock')  # MB52 – Material Status (Active/Inactive) &Stock levels
    sl = fields.Char('Active/Inactive')

    task_list = fields.Char('Task List')  # Z4190 – Task List Details related to Material

    last_wo_used_date = fields.Char('Last WO used date')
    stores_min_max = fields.Char('Stores Min-Max Levels')  # Z5001 – Stores Min-Max Levels

    supplier_details = fields.Char('Supplier Details')  # ME2M – Supplier Details
    supplier_part_no = fields.Char('Supplier Part No')
    quotes_price = fields.Char('Quotes Price')  # MM60 – Unit Price of the Material
    supply_lead_time = fields.Char('Supply Lead Time')  # Z7116 – Material Supply Lead Time
    long_text = fields.Char('Long Text')  # Sphera (Spares Finder) - Material Long Text

    material_name = fields.Char('Material Name')  # Z7134 – HERS record – Mfg./Mfg. Part No details
    manufacturer_name = fields.Char('Manufacturer Name')
    manufacturer_part_no = fields.Char('Manufacturer Part No')
    manufacturer_description = fields.Char('Material Description')

    liner = fields.Many2one("product.liner_taxonomy")

    bom_status = fields.Selection(string='BoM Status',
                                  selection=[('approved_to_add', 'Approved to Add'),  # 0
                                             ('approved_to_update', 'Approved to Update'),  # 2
                                             ('approved_to_delete', 'Approved to Delete'),  # 3
                                             ('decline_to_delete', 'Decline To Delete'),  # 4
                                             ('no_action', 'No Action Required'),  # 5
                                             ('approved', 'Approved'),  # 6
                                             ('decline', 'Decline'),  # 7
                                             ])

    is_duplicate = fields.Boolean("SAP Material Duplicate")
    diff_r_code = fields.Char('Rtio Liner Code (Count > 1) with NEW and Material number(2 or 4 series)')

    def taxonomy_bom_structure_same_material_with_differnt_rio_code(self):
        q = """SELECT rio_code FROM taxonomy_bom_structure GROUP BY rio_code 
                HAVING COUNT(*) > 1 AND COUNT(CASE WHEN material_number LIKE 'NEW%' THEN 1 END) > 0 
                AND COUNT(CASE WHEN material_number LIKE '2%' OR material_number LIKE '4%' THEN 1 END) > 0;"""
        self.env.cr.execute(q)
        result = self.env.cr.fetchall()
        rio_code = [r[0] for r in result]
        if len(rio_code) == 1:
            rio_code_tuple = f"('{rio_code[0]}')"
        else:
            rio_code_tuple = tuple(rio_code)

        q2 = f"UPDATE taxonomy_bom_structure SET diff_r_code = False;"
        self.env.cr.execute(q2)

        if rio_code_tuple:
            q3 = f"""UPDATE taxonomy_bom_structure SET diff_r_code = 'Rtio Liner Code (Count > 1) with NEW and Material number(2 or 4 series)'
                        where rio_code IN {rio_code_tuple};"""
            self.env.cr.execute(q3)

    def re_update_BoM(self):
        boms = self.env['taxonomy.bom_structure'].search([('liner_id', '!=', False)])
        for bom in boms:
            lin_ids = bom.liner_id.split(",")
            liner_ids = [line.strip() for line in lin_ids]
            liner_taxonomy = self.env['product.liner_taxonomy'].search([('wearlinerid', 'in', list(liner_ids))])
            vals = self.add_liner_info({}, liner_taxonomy)
            bom.write(vals)

    def button_bom_status_change(self):
        result = self.env["ir.actions.act_window"]._for_xml_id(
            "rio_cyeims.wizard_bom_status_move_act_window"
        )
        result["context"] = dict(self.env.context)
        return result

    def button_generate_spares_finder(self):
        result = self.env["ir.actions.act_window"]._for_xml_id(
            "rio_cyeims.wizard_bom_spares_finder_act_window"
        )
        result["context"] = dict(self.env.context)
        return result

    def get_material_status(self, liners, ih01_flocs):
        liner_flocs = [liner.floc_id.name for liner in liners]
        if ih01_flocs in liner_flocs:
            return 'match'
        else:
            return 'existing_flocs_mismatch'

    def get_flocs_as_str(self, liners):
        separator = ", "
        vals = []
        for liner in liners:
            vals.append(liner.floc_id.name)
        vals = list(set(vals))
        element_str = separator.join(vals)
        return element_str

    @api.model
    def get_string_of_fields(self, element, liners, separator=None):
        separator = separator if separator else ", "
        vals = []
        for liner in liners:
            value = getattr(liner, element)
            if isinstance(value, str):
                value = value if value else ''
            if isinstance(value, int):
                value = str(value) if value else ''
            #value = value if isinstance(value, str) else ''
            vals.append(value)
        vals = list(set(vals))

        element_str = separator.join(vals)
        return element_str

    def get_str_val(self, val):
        val = val if isinstance(val, str) else ''
        return val

    def add_liner_info(self, vals, liner_taxonomy):

        vals['material_where_used'] = self.get_flocs_as_str(liner_taxonomy)
        vals['rio_code'] = self.get_string_of_fields('rtio_liner_code', liner_taxonomy)
        vals['liner_id'] = self.get_string_of_fields('wearlinerid', liner_taxonomy)
        vals['detail_drawing'] = self.get_string_of_fields('dwg_linerdetail', liner_taxonomy)
        vals['layout_drawing_no'] = self.get_string_of_fields('dwg_markingplan', liner_taxonomy)
        vals['item_no'] = self.get_string_of_fields('itemno', liner_taxonomy)
        vals['std_liner'] = self.get_string_of_fields('standard_liner', liner_taxonomy)
        vals['sap_long_text'] = self.get_string_of_fields('sap_long_text', liner_taxonomy,'\n\n')
        vals['sap_short_text'] = self.get_string_of_fields('sap_short_text', liner_taxonomy)
        vals['material_qty'] = self.get_string_of_fields('liner_detail_qty', liner_taxonomy)
        vals['material_description'] = self.get_string_of_fields('sap_short_text', liner_taxonomy)
        vals['material_unit'] = 'EA'
        return vals

    def add_sap_t_code_vals(self, vals, material, flocs):
        # MB51
        mb51 = self.env['mb51.sample_customized'].search([('material', '=', material)], limit=1)
        if mb51:
            vals['plant'] = mb51.plant
            vals['last_wo_used'] = mb51.order
            vals['goods_movement'] = f'{mb51.movement_type} - {mb51.movement_type_text}'

        mb52 = self.env['mb52.sample_customized'].search([('material', '=', material)], limit=1)
        if mb52:
            vals['current_stock'] = mb52.unrestricted
            vals['sl'] = mb52.df_stor_loc_level

        if flocs:
            z4190 = self.env['z4190.sample_customized'].search([('component', '=', material),
                                                                ('functional_location', '=', flocs)], limit=1)
            if z4190:
                vals['task_list'] = f'{z4190.group} - {z4190.grpcountr}'

        iw38 = self.env['iw38.sample_customized'].search([('material', '=', material)], limit=1)
        if iw38:
            vals['last_wo_used_date'] = iw38.bas_start_date

        z5001 = self.env['z5001.sample_customized'].search([('material', '=', material)], limit=1)
        if z5001:
            vals['stores_min_max'] = f'{z5001.minimum_stock} - {z5001.maximum_stock}'

        me2m = self.env['me2m.sample_customized'].search([('material', '=', material)], limit=1)
        if me2m:
            vals['supplier_details'] = me2m.vendor_supplying_plant
            vals['supplier_part_no'] = ''

        mm60 = self.env['mm60.sample_customized'].search([('material', '=', material)], limit=1)
        if mm60:
            vals['quotes_price'] = mm60.price

        z7116 = self.env['z7116.sample_customized'].search([('material', '=', material)], limit=1)
        if z7116:
            vals['supply_lead_time'] = z7116.ltime

        spare = self.env['spare.finder_customized'].search([('material', '=', material)], limit=1)
        if spare:
            vals['long_text'] = spare.long_description

        z7134 = self.env['z7134.sample_customized'].search([('int_no', '=', material)], limit=1)
        if z7134:
            vals['material_name'] = ''
            vals['manufacturer_name'] = z7134.mfr
            vals['manufacturer_part_no'] = z7134.mpn
            vals['manufacturer_description'] = z7134.internal_material_description

        return vals

    def add_parent_chut_flocs(self, site_id):
        _logger.info(f"add_parent_chut_flocs: Started")
        ih01_data = self.env['ih01.bom_data'].read_group([('floc', 'ilike', 'CHU'), ('parent_floc', 'ilike', 'CHU'),
                                                          ('site', '=', site_id)],
                                                         ['floc', 'ct', 'pma', 'parent_floc'], ['parent_floc'])
        for iho1_d in ih01_data:
            iho1 = self.env['ih01.bom_data'].search(iho1_d['__domain'], limit=1)
            vals = {'pma_description': self.get_str_val(iho1.pma_desc), 'site': iho1.site.id,
                    'package_unit': self.get_str_val(iho1.pu),
                    'functional_location': self.get_str_val(iho1.parent_floc),
                    'floc_description': self.get_str_val(iho1.parent_floc_desc),
                    'parent_pma': self.get_str_val(iho1.parent_pma), 'new_existing_material': 'undefined'}
            self.env['taxonomy.bom_structure'].create(vals)
        _logger.info(f"add_parent_chut_flocs: Completed")

    def create_bom_structure(self, iho1):
        vals = {'pma_description': self.get_str_val(iho1.pma_desc),
                'material_description': self.get_str_val(iho1.material_desc),
                'material_qty': self.get_str_val(iho1.material_qty),
                'material_unit': self.get_str_val(iho1.material_unit), 'site': iho1.site.id,
                'package_unit': self.get_str_val(iho1.pu), 'functional_location': self.get_str_val(iho1.floc),
                'floc_description': self.get_str_val(iho1.floc_desc),
                'material_number': self.get_str_val(iho1.material), 'parent_pma': self.get_str_val(iho1.parent_pma),
                'pma': self.get_str_val(iho1.pma), 'ct': self.get_str_val(iho1.ct)}

        bom = self.env['taxonomy.bom_structure'].search([('package_unit', '=', self.get_str_val(iho1.pu)),
                                                         ('functional_location', '=', self.get_str_val(iho1.floc)),
                                                         ('material_number', '=', self.get_str_val(iho1.material)),
                                                         ('pma', '=', self.get_str_val(iho1.pma)),
                                                         ('ct', '=', self.get_str_val(iho1.ct)),
                                                         ])
        if iho1.material:
            liner_taxonomy = self.env['product.liner_taxonomy'].search(
                [('sap_material', '=', self.get_str_val(iho1.material)), ('site', '=', iho1.site.id)])
            if liner_taxonomy:

                status = self.get_material_status(liner_taxonomy, self.get_str_val(iho1.floc))
                if status == 'match':
                    liner_flocs = [liner for liner in liner_taxonomy if
                                   liner.floc_id.name == self.get_str_val(iho1.floc)]
                    vals = self.add_liner_info(vals, liner_flocs)
                elif status == 'existing_flocs_mismatch':
                    vals['material_where_used'] = self.get_flocs_as_str(liner_taxonomy)
                vals['new_existing_material'] = status
                liner_taxonomy.write({'is_exist_sap': True})
            else:
                vals['new_existing_material'] = 'existing_no_match'

            vals = self.add_sap_t_code_vals(vals, iho1.material, iho1.floc)
        else:
            vals['new_existing_material'] = 'undefined'

        if bom:
            vals['is_duplicate'] = True

        self.env['taxonomy.bom_structure'].create(vals)

    def update_other_fields_bom_structure(self, site_id):
        offset = 0
        limit = 500
        while True:
            res_data = self.env['ih01.bom_data'].search([('floc', 'ilike', 'CHU'), ('site', '=', site_id)],
                                                        offset=offset, limit=limit)
            if not res_data:
                break

            for iho1 in res_data:
                self.create_bom_structure(iho1)

            offset += limit
            self._cr.commit()
            _logger.info(f"update_other_fields_bom_structure: {offset}")

        self.add_parent_chut_flocs(site_id)
        self.update_existing_flocs_mismatch_fields_bom_from_liner(site_id)

    def update_existing_flocs_mismatch_fields_bom_from_liner(self, site_id=None):
        if site_id:
            bom_existing_flocs_mismatch = self.search(
                [('new_existing_material', '=', 'existing_flocs_mismatch'),
                 ('site', '=', site_id)])
        else:
            bom_existing_flocs_mismatch = self.search(
                [('new_existing_material', '=', 'existing_flocs_mismatch')])

        for bom in bom_existing_flocs_mismatch:
            if not bom.material_where_used:
                continue

            flocs = bom.material_where_used.split(",")
            liners = self.env['product.liner_taxonomy'].search([
                ('sap_material', '=', bom.material_number),
                ('floc_id', 'in', flocs)])
            if liners:
                def get_joined_unique(field_name):
                    values = list(set(filter(None, liners.mapped(field_name))))
                    return ','.join(values) if len(values) > 1 else (values[0] if values else False)

                item_no_val = get_joined_unique('itemno')
                layout_drawing_no_val = get_joined_unique('dwg_markingplan')
                detail_drawing_val = get_joined_unique('dwg_linerdetail')
                vals = {
                    'item_no': item_no_val,
                    'layout_drawing_no': layout_drawing_no_val,
                    'detail_drawing': detail_drawing_val,
                }
                bom.write(vals)

    def is_special_case(self,domain):
        offset = 0
        limit = 1000
        while True:
            res_data = self.env['ih01.bom_data'].search(domain,
                                                        offset=offset, limit=limit)
            if not res_data:
                break

            for iho1 in res_data:
                self.create_bom_structure(iho1)

            offset += limit
            self._cr.commit()

    def generate_pma(self, flocs):
        bom_data = self.env['taxonomy.bom_structure'].read_group(
            [('functional_location', '=', flocs)], ['pma'], ['pma'])
        pma_list = list(set(item['pma'] for item in bom_data if item['pma'] != ''))
        pma_list = [x for x in pma_list if x is not False]

        if len(pma_list) == 1:
            return pma_list.pop()

        for pma in pma_list:
            if 'NEWPMA' in pma:
                return pma
        return self.env['ir.sequence'].next_by_code('rio.new.pma')

    def generate_ct(self, flocs):
        bom_data = self.env['taxonomy.bom_structure'].read_group(
            [('functional_location', '=', flocs)], ['ct'], ['ct'])
        ct_list = list(set(item['ct'] for item in bom_data if item['ct'] != ''))
        ct_list = [x for x in ct_list if x is not False]
        if len(ct_list) == 1:
            return ct_list.pop()

        for ct in ct_list:
            if 'NEWCT' in ct:
                return ct
        return self.env['ir.sequence'].next_by_code('rio.new.ct')

    def generate_new_material(self):
        return self.env['ir.sequence'].next_by_code('rio.new.material_number')

    def update_bom_from_liner(self, site):
        offset = 0
        limit = 1000

        while True:
            liners = self.env['product.liner_taxonomy'].search([('site', '=', site)],
                                                               offset=offset, limit=limit)
            if not liners:
                break
            for liner in liners:

                flocs = liner.floc_id.name
                material_number = liner.sap_material

                if material_number != 'NA':
                    bom_liner = self.env['taxonomy.bom_structure'].search([('material_number', '=', liner.sap_material),
                                                                           ('functional_location', '=', flocs)])
                    if bom_liner:
                        continue
                    bom_liner = self.env['taxonomy.bom_structure'].search([('functional_location', '=', flocs),
                                                                           ('ct', '=', ''),
                                                                           ('pma', '=', '')])
                    if bom_liner:
                        vals = {
                            'material_number': liner.sap_material,
                            'pma': self.generate_pma(flocs),
                            'ct': self.generate_ct(flocs),
                            'liner': liner.id,
                            'new_existing_material': 'add'}
                    else:
                        vals = {'site': liner.site.id,
                                'package_unit': liner.pu_id.name,
                                'functional_location': flocs,
                                'floc_description': liner.floc_desc,
                                'material_number': liner.sap_material,
                                'pma': self.generate_pma(flocs),
                                'ct': self.generate_ct(flocs),
                                'liner': liner.id,
                                'new_existing_material': 'add'}
                        bom_liner = self.env['taxonomy.bom_structure'].search([('liner', '=', liner.id)])
                    vals = self.add_sap_t_code_vals(vals, material_number, flocs)
                else:

                    vals = {'site': liner.site.id,
                            'package_unit': liner.pu_id.name,
                            'functional_location': flocs,
                            'floc_description': liner.floc_desc,
                            'material_number': self.generate_new_material(),
                            'pma': self.generate_pma(flocs),
                            'ct': self.generate_ct(flocs),
                            'liner': liner.id,
                            'new_existing_material': 'new_material'}
                    bom_liner = self.env['taxonomy.bom_structure'].search([('liner', '=', liner.id)])

                vals = self.add_liner_info(vals, [liner])

                if bom_liner:
                    bom_liner.write(vals)
                else:
                    self.env['taxonomy.bom_structure'].create(vals)
            offset += limit
            self._cr.commit()
            _logger.info(f"update_bom_from_liner: {offset}")

    def re_change_create_new_material_in_sap(self):

        bom_data = self.env['taxonomy.bom_structure'].read_group(
            [('new_existing_material', '=', 'new_material'), ('material_number', 'ilike', 'NEWMAT%')], ['rio_code'],
            ['rio_code'])
        # same_riocode_material
        for bom in bom_data:
            bom_with_material = self.env['taxonomy.bom_structure'].read_group(
                [('rio_code', '=', bom['rio_code']), ('material_number', 'not ilike', 'NEWMAT%')], ['rio_code'],
                ['rio_code'])
            if bom_with_material:
                boms = self.env['taxonomy.bom_structure'].search(bom['__domain'])
                boms.write({'new_existing_material': 'same_riocode_material'})
